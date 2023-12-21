import tkinter as tk
from tkinter import ttk
from datetime import datetime
from tkinter import messagebox

import openpyxl
import re
def is_valid_name(name, min_length=2, max_length=20, allowed_characters=r"^[a-zA-Z '-]+$"):
    if not name:
        return False

    if not re.match(allowed_characters, name):
        return False

    if not min_length <= len(name) <= max_length:
        return False

    return True

def is_valid_phone(phone):
    phone_pattern = re.compile(r"^\d{10}$")
    return bool(re.match(phone_pattern, phone))

class Order:
    def __init__(self, order_id, customer_name, customer_phone, order_date, cake_type, cake_amount):
        self.order_id = order_id
        self.customer_name = customer_name
        self.customer_phone = customer_phone
        self.order_date = order_date
        self.cake_type = cake_type
        self.cake_amount = cake_amount

class BakeryManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Bakery Management System")

        self.orders = load_orders_from_excel()

        # Add Order Frame
        add_order_frame = ttk.LabelFrame(root, text="Add Order")
        add_order_frame.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        ttk.Label(add_order_frame, text="Customer Name:").grid(row=0, column=0, padx=5, pady=5)
        self.customer_name_entry = ttk.Entry(add_order_frame)
        self.customer_name_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(add_order_frame, text="Customer Phone:").grid(row=1, column=0, padx=5, pady=5)
        self.customer_phone_entry = ttk.Entry(add_order_frame)
        self.customer_phone_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(add_order_frame, text="Cake Choice:").grid(row=2, column=0, padx=5, pady=5)
        self.cake_choice_combobox = ttk.Combobox(add_order_frame, values=[
            "Chocolate Fudge Cake", "Red Velvet Cake", "Carrot Cake", "Vanilla Bean Cheesecake",
            "Banana Bread Cake", "Lemon Pound Cake", "Angel Food Cake", "Coffee Cake",
            "Black Forest Cake", "Hummingbird Cake"
        ])
        self.cake_choice_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.cake_choice_combobox.set("Chocolate Fudge Cake")

        ttk.Label(add_order_frame, text="Cake Amount:").grid(row=3, column=0, padx=5, pady=5)
        self.cake_amount_entry = ttk.Entry(add_order_frame)
        self.cake_amount_entry.grid(row=3, column=1, padx=5, pady=5)

        ttk.Button(add_order_frame, text="Add Order", command=self.add_order).grid(row=4, column=0, columnspan=2, pady=10)

        # Search Order Frame
        search_order_frame = ttk.LabelFrame(root, text="Search Order")
        search_order_frame.grid(row=0, column=1, padx=10, pady=10, sticky="e")

        ttk.Label(search_order_frame, text="Order ID:").grid(row=0, column=0, padx=5, pady=5)
        self.search_order_id_entry = ttk.Entry(search_order_frame)
        self.search_order_id_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Button(search_order_frame, text="Search Order", command=self.search_order).grid(row=1, column=0, columnspan=2, pady=10)

        # Excel Update Frame
        update_excel_frame = ttk.LabelFrame(root, text="Update Excel")
        update_excel_frame.grid(row=0, column=2, padx=10, pady=10, sticky="e")

        ttk.Button(update_excel_frame, text="Update Excel", command=self.update_excel).grid(row=0, column=0, pady=10)

    def add_order(self):
        customer_name = self.customer_name_entry.get()
        if not is_valid_name(customer_name):
            self.show_message("Error", "Can't proceed, the name is not valid.")
            return

        customer_phone = self.customer_phone_entry.get()
        if not is_valid_phone(customer_phone):
            self.show_message("Error", "Invalid phone number. Please enter a 10-digit number.")
            return

        cake_choice = self.cake_choice_combobox.get()
        cake_amount = self.cake_amount_entry.get()

        if not cake_amount.isdigit():
            self.show_message("Error", "Invalid cake amount. Please enter a valid number.")
            return

        order_date = datetime.utcnow().strftime("%Y-%m-%d")
        order_id = len(self.orders) + 1

        new_order = Order(order_id, customer_name, customer_phone, order_date, cake_choice, int(cake_amount))
        self.orders.append(new_order)

        self.show_message("Success", f"Order added successfully. Order ID: {new_order.order_id}")

        self.clear_add_order_entries()
        update_excel_file(self.orders)

    def search_order(self):
        order_id = self.search_order_id_entry.get()

        if not order_id.isdigit():
            self.show_message("Error", "Invalid order ID. Please enter a valid number.")
            return

        order_id = int(order_id)
        found_order = None

        for order in self.orders:
            if order.order_id == order_id:
                found_order = order
                break

        if found_order:
            self.show_message("Order Details", f"Order ID: {found_order.order_id}\n"
                                               f"Customer Name: {found_order.customer_name}\n"
                                               f"Customer Phone: {found_order.customer_phone}\n"
                                               f"Order Date: {found_order.order_date}\n"
                                               f"Cake Type: {found_order.cake_type}\n"
                                               f"Cake Amount: {found_order.cake_amount}")
        else:
            self.show_message("Error", "Order not found.")

    def update_excel(self):
        update_excel_file(self.orders)
        self.show_message("Success", "Excel file updated successfully.")

    def clear_add_order_entries(self):
        self.customer_name_entry.delete(0, tk.END)
        self.customer_phone_entry.delete(0, tk.END)
        self.cake_choice_combobox.set("Chocolate Fudge Cake")
        self.cake_amount_entry.delete(0, tk.END)

    def show_message(self, title, message):
        tk.messagebox.showinfo(title, message)

def load_orders_from_excel():
    orders = []

    try:
        workbook = openpyxl.load_workbook("bakery_orders.xlsx")
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row) == 6:
                order = Order(row[0], row[1], row[2], row[3], row[4], row[5])
                orders.append(order)
            else:
                print("Skipping invalid row in the Excel file.")

    except FileNotFoundError:
        pass

    return orders

def update_excel_file(orders):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.append(["Order ID", "Customer Name", "Customer Phone", "Order Date", "Cake Type", "Cake Amount"])

    for order in orders:
        worksheet.append([order.order_id, order.customer_name, order.customer_phone,
                          order.order_date, order.cake_type, order.cake_amount])

    workbook.save("bakery_orders.xlsx")

if __name__ == "__main__":
    root = tk.Tk()
    app = BakeryManagementApp(root)
    root.mainloop()
