import openpyxl
import re
from datetime import datetime


def is_valid_name(name, min_length=2, max_length=20, allowed_characters=r"^[a-zA-Z '-]+$"):
    """Checks if a name is valid based on specified criteria.

    Args:
        name (str): The name to check.
        min_length (int, optional): Minimum allowed length for the name. Defaults to 2.
        max_length (int, optional): Maximum allowed length for the name. Defaults to 30.
        allowed_characters (str, optional): A regular expression pattern defining valid characters.
            Defaults to letters, spaces, apostrophes, and hyphens.

    Returns:
        bool: True if the name is valid, False otherwise.
    """
    if not name:
        return False  # Empty names are invalid

    if not re.match(allowed_characters, name):
        return False  # Name contains invalid characters

    if not min_length <= len(name) <= max_length:
        return False  # Name length is outside the allowed range

    # Additional validation rules you can add here...

    return True


def is_valid_phone(phone):
    phone_pattern = re.compile(r"^\d{10}$")  # Assuming a 10-digit phone number
    return bool(re.match(phone_pattern, phone))

def search_order_by_id(orders, order_id):
    for order in orders:
        if order.order_id == order_id:
            print(f"Order ID: {order.order_id}")
            print(f"Customer Name: {order.customer_name}")
            print(f"Customer Phone: {order.customer_phone}")
            print(f"Order Date: {order.order_date}")
            print(f"Cake Type: {order.cake_type}")
            print(f"Cake Amount: {order.cake_amount}")
            return

    print("Order not found.")


    

# Structure to represent an order
class Order:
    def __init__(self, order_id, customer_name, customer_phone, order_date, cake_type, cake_amount):
        self.order_id = order_id
        self.customer_name = customer_name
        self.customer_phone = customer_phone
        self.order_date = order_date
        self.cake_type = cake_type
        self.cake_amount = cake_amount


# Function to display the cake menu and get user choice
def get_cake_choice():
    print("Cake Menu:")
    print("1. Chocolate Fudge Cake")
    print("2. Red Velvet Cake")
    print("3. Carrot Cake")
    print("4. Vanilla Bean Cheesecake")
    print("5. Banana Bread Cake")
    print("6. Lemon Pound Cake")
    print("7. Angel Food Cake")
    print("8. Coffee Cake")
    print("9. Black Forest Cake")
    print("10. Hummingbird Cake")

    while True:
        choice = input("Enter the number corresponding to your cake choice: ")
        if choice.isdigit() and 1 <= int(choice) <= 10:
            return int(choice)
        else:
            print("Invalid choice. Please enter a valid number.")


def get_customer_phone():
    while True:
        phone = input("Enter customer phone number (10 digits): ")
        if is_valid_phone(phone):
            return phone
        else:
            print("Invalid phone number. Please enter a 10-digit number.")


def add_order(orders):
    customer_name = input("Enter customer name: ")
    if not is_valid_name(customer_name):
        print("Can't proceed, the name is not valid.")
        return

    customer_phone = get_customer_phone()

    order_date = datetime.utcnow().strftime("%Y-%m-%d")
    cake_choice = get_cake_choice()
    cake_types = ["Chocolate Fudge Cake", "Red Velvet Cake", "Carrot Cake", "Vanilla Bean Cheesecake",
                  "Banana Bread Cake", "Lemon Pound Cake", "Angel Food Cake", "Coffee Cake",
                  "Black Forest Cake", "Hummingbird Cake"]
    cake_type = cake_types[cake_choice - 1]

    cake_amount = int(input("Enter cake amount: "))

    order_id = len(orders) + 1
    new_order = Order(order_id, customer_name, customer_phone, order_date, cake_type, cake_amount)

    orders.append(new_order)
    print(f"Order added successfully. Order ID: {new_order.order_id}")
    update_excel_file(orders)


# Function to update Excel file with order details
def update_excel_file(orders):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write headers to the Excel file
    worksheet.append(["Order ID", "Customer Name", "Customer Phone", "Order Date", "Cake Type", "Cake Amount"])

    # Write order details to the Excel file
    for order in orders:
        worksheet.append([order.order_id, order.customer_name, order.customer_phone,
                          order.order_date, order.cake_type, order.cake_amount])

    # Save the Excel file
    workbook.save("bakery_orders.xlsx")

    print("Excel file updated successfully.")


# Function to load existing orders from Excel file (if any)
def load_orders_from_excel():
    orders = []

    try:
        workbook = openpyxl.load_workbook("bakery_orders.xlsx")
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row) == 6:  # Make sure there are six values in the row
                order = Order(row[0], row[1], row[2], row[3], row[4], row[5])
                orders.append(order)
            else:
                print("Skipping invalid row in the Excel file.")

    except FileNotFoundError:
        pass

    return orders


# Main function
def main():
    orders = load_orders_from_excel()

    while True:
        print("\nBakery Management System")
        print("1. Add Order")
        print("2. Search Order")  # You can add more functionality as needed
        print("3. Update Order")
        print("4. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            add_order(orders)
        elif choice == "4":
            print("Exiting the Bakery Management System. Goodbye!")
            break
        elif choice=='2':
            search_order_by_id(orders,int(input("Enter the order id : ")))

        else:
            print("Invalid choice. Please enter a valid option.")


if __name__ == "__main__":
    main()
